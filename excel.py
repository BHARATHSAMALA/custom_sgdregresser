import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Color, PatternFill, Font, GradientFill, Alignment
from openpyxl.utils import get_column_letter

# Load the workbook
wb = load_workbook('example.xlsx')

# Select the sheet
sheet = wb.active

# Get the min and max values for each feature column
feature_cols = [col for col in sheet.iter_cols(min_row=1, max_row=1)]
for col in feature_cols:
    values = [cell.value for cell in col[1:]]
    col_min = min(values)
    col_max = max(values)
    
    # Define the color scale
    red = Color(rgb='FF0000')
    yellow = Color(rgb='FFFF00')
    green = Color(rgb='00FF00')
    scale = GradientFill(stop=[ 
        GradientStop(position=0, color=red), 
        GradientStop(position=0.5, color=yellow), 
        GradientStop(position=1, color=green) 
    ])

    # Color code each cell in the column based on its value
    for cell in col:
        if cell.row == 1:
            continue  # Skip the header row
        value = cell.value
        if value == col_min:
            cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        elif value == col_max:
            cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        else:
            position = (value - col_min) / (col_max - col_min)
            cell.fill = scale[position]

# Save the workbook
wb.save('example_colored.xlsx')
